

from openpyxl import Workbook
wb=Workbook()
#grap the active worksheet
ws=wb.active

# data can be associated directly to the cell
ws['A1']=42

#rows can also be appended
ws.append([1,2,3])

# python type will automatically to converted
import datetime
ws['A2']=datetime.datetime.now()

#save the file
wb.save('sample.xlsx')
